from flask import Blueprint, render_template, request, redirect, jsonify, send_file
from models import get_db, pool
from werkzeug.utils import secure_filename
from flask import session
from datetime import datetime
from flask import jsonify
import json
import os
import uuid
import re
import requests
from difflib import SequenceMatcher
from io import BytesIO
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

UPLOAD_DIR = os.path.join(
    "static",
    "uploads",
    "items"
)

os.makedirs(UPLOAD_DIR, exist_ok=True)

items_bp = Blueprint("items", __name__)


@items_bp.route("/items")
def items():
    conn = get_db()
    
    cur = conn.cursor()
    
    cur.execute("""
    SELECT 
        items.*,
        (SELECT image FROM item_images 
         WHERE item_id = items.id 
         LIMIT 1) as image
    FROM items
    WHERE items.company_id = %s
    ORDER BY items.id DESC
    LIMIT 50
    """, (session.get("company_id"),))
    items = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) AS total
        FROM items
        WHERE company_id = %s
    """, (session.get("company_id"),))
    catalog_total = cur.fetchone()["total"]

    cur.execute("""
        SELECT *
        FROM categories
        WHERE company_id = %s
        ORDER BY name
    """, (session.get("company_id"),))
    categories = cur.fetchall()

    pool.putconn(conn)
    return render_template(
        "items.html",
        items=items,
        categories=categories,
        catalog_total=catalog_total,
        catalog_page_size=50,
    )


def _catalog_item_payload(item):
    """Convert a database row into a compact JSON-safe catalog item."""
    data = dict(item)
    for field in (
        "retail_price", "purchase_price", "wholesale_price", "quantity",
        "discount_percent", "markup_percent",
    ):
        value = data.get(field)
        if isinstance(value, Decimal):
            data[field] = float(value)
    if isinstance(data.get("is_marked"), Decimal):
        data["is_marked"] = bool(data["is_marked"])
    return data


@items_bp.route("/api/catalog/items")
def api_catalog_items():
    company_id = session.get("company_id")
    query = (request.args.get("q") or "").strip()
    item_type = (request.args.get("type") or "all").strip().lower()
    category = (request.args.get("category") or "all").strip()

    try:
        page = max(1, int(request.args.get("page", 1)))
    except (TypeError, ValueError):
        page = 1

    try:
        limit = min(100, max(1, int(request.args.get("limit", 50))))
    except (TypeError, ValueError):
        limit = 50

    where = ["items.company_id = %s"]
    params = [company_id]

    if item_type in ("product", "service"):
        where.append("COALESCE(items.item_type, 'product') = %s")
        params.append(item_type)

    if category and category.lower() != "all":
        where.append("LOWER(COALESCE(items.category, '')) = LOWER(%s)")
        params.append(category)

    if query:
        pattern = f"%{query}%"
        where.append("""(
            items.name ILIKE %s
            OR COALESCE(items.barcode, '') ILIKE %s
            OR COALESCE(items.gtin, '') ILIKE %s
            OR COALESCE(items.ntin, '') ILIKE %s
        )""")
        params.extend([pattern, pattern, pattern, pattern])

    where_sql = " AND ".join(where)
    offset = (page - 1) * limit
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute(
            f"SELECT COUNT(*) AS total FROM items WHERE {where_sql}",
            tuple(params),
        )
        total = cur.fetchone()["total"]

        cur.execute(f"""
            SELECT
                items.*,
                (SELECT image FROM item_images
                 WHERE item_id = items.id
                 LIMIT 1) AS image
            FROM items
            WHERE {where_sql}
            ORDER BY items.id DESC
            LIMIT %s OFFSET %s
        """, tuple(params + [limit, offset]))
        rows = [_catalog_item_payload(row) for row in cur.fetchall()]
    finally:
        pool.putconn(conn)

    return jsonify({
        "success": True,
        "items": rows,
        "page": page,
        "limit": limit,
        "total": total,
        "has_more": offset + len(rows) < total,
    })

@items_bp.route("/items/add", methods=["GET", "POST"])
def add_item():

    conn = get_db()
    
    cur = conn.cursor()

    if request.method == "POST":

        company_id = session.get("company_id")

        cur.execute("""
            INSERT INTO items (
                name,
                category,
                unit,
                description,
                retail_price,
                wholesale_price,
                purchase_price,
                discount_percent,
                barcode,
                gtin,
                ntin,
                is_marked,
                item_type,
                service_sale_mode,
                company_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            request.form["name"],
            request.form["category"],
            request.form.get("unit"),
            request.form.get("description"),
            float(request.form.get("retail_price") or 0),
            float(request.form.get("wholesale_price") or 0),
            float(request.form.get("purchase_price") or 0),
            int(request.form.get("discount_percent") or 0),
            request.form.get("barcode"),
            request.form.get("gtin"),
            request.form.get("ntin"),
            request.form.get("is_marked") == "1" if request.form.get("item_type", "product") == "product" else False,
            request.form.get("item_type", "product"),
            (request.form.get("service_sale_mode") or "order") if request.form.get("item_type", "product") == "service" else None,
            company_id
        ))

        item_id = cur.fetchone()["id"]

        # 🔥 загрузка картинок

        images = request.files.getlist("images")

        for image in images:

            if image and image.filename:

                filename = secure_filename(image.filename)

                filename = (
                    f"{uuid.uuid4().hex}_{filename}"
                )

                save_path = os.path.join(
                    UPLOAD_DIR,
                    filename
                )

                image.save(save_path)

                image_path = (
                    "/" + save_path.replace("\\", "/")
                )

                cur.execute("""
                    INSERT INTO item_images
                    (
                        item_id,
                        image
                    )
                    VALUES (%s, %s)
                """, (
                    item_id,
                    image_path
                ))

        conn.commit()

        pool.putconn(conn)

        return redirect("/items")

    # категории
    cur.execute("""
        SELECT *
        FROM categories
        WHERE company_id = %s
        ORDER BY name
    """, (
        session.get("company_id"),
    ))
    
    categories = cur.fetchall()

    pool.putconn(conn)

    return render_template(
        "item_form.html",
        categories=categories
    )
    
@items_bp.route("/items/<int:item_id>/edit", methods=["GET", "POST"])
def edit_item(item_id):
    conn = get_db()
    
    cur = conn.cursor()

    if request.method == "POST":
        cur.execute("""
            UPDATE items
            SET 
                name = %s,
                category = %s,
                unit = %s,
                description = %s,
                retail_price = %s,
                wholesale_price = %s,
                purchase_price = %s,
                discount_percent = %s,
                barcode = %s,
                gtin = %s,
                ntin = %s,
                is_marked = %s,
                item_type = %s,
                service_sale_mode = %s
            WHERE id = %s AND company_id = %s
        """, (
            request.form["name"],
            request.form["category"],
            request.form.get("unit"),
            request.form.get("description"),
            float(request.form.get("retail_price") or 0),
            float(request.form.get("wholesale_price") or 0),
            float(request.form.get("purchase_price") or 0),
            int(request.form.get("discount_percent") or 0),
            request.form.get("barcode"),
            request.form.get("gtin"),
            request.form.get("ntin") if request.form.get("item_type", "product") == "product" else "",
            request.form.get("is_marked") == "1" if request.form.get("item_type", "product") == "product" else False,
            request.form.get("item_type", "product"),
            (request.form.get("service_sale_mode") or "order") if request.form.get("item_type", "product") == "service" else None,
            item_id,
            session.get("company_id")
        ))
        
        # Новое изображение при редактировании
        images = request.files.getlist("images")
        new_images = [
            image for image in images
            if image and image.filename
        ]

        if new_images:
            # Получаем старые изображения
            cur.execute("""
                SELECT image
                FROM item_images
                WHERE item_id = %s
            """, (item_id,))

            old_images = cur.fetchall()

            # Удаляем старые записи из БД
            cur.execute("""
                DELETE FROM item_images
                WHERE item_id = %s
            """, (item_id,))

            # Удаляем старые файлы
            for old_image in old_images:
                old_path = old_image["image"]

                if old_path:
                    file_path = old_path.lstrip("/")

                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                        except OSError:
                            pass

            # Сохраняем новое изображение
            for image in new_images:
                filename = secure_filename(image.filename)

                filename = f"{uuid.uuid4().hex}_{filename}"

                save_path = os.path.join(
                    UPLOAD_DIR,
                    filename
                )

                image.save(save_path)

                image_path = "/" + save_path.replace("\\", "/")

                cur.execute("""
                    INSERT INTO item_images (
                        item_id,
                        image
                    )
                    VALUES (%s, %s)
                """, (
                    item_id,
                    image_path
                ))

        conn.commit()
        pool.putconn(conn)
        return redirect("/items")

    cur.execute(
        "SELECT * FROM items WHERE id = %s AND company_id = %s",
        (item_id, session.get("company_id"))
    )

    item = cur.fetchone()

    conn = get_db()
    
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM categories
        WHERE company_id = %s
        ORDER BY name
    """, (
        session.get("company_id"),
    ))

    categories = cur.fetchall()

    pool.putconn(conn)

    return render_template(
        "item_form.html",
        item=item,
        categories=categories
    )
    
@items_bp.route("/items/<int:item_id>/delete")
def delete_item(item_id):
    conn = get_db()
    
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM items WHERE id = %s AND company_id = %s",
        (item_id, session.get("company_id"))
    )

    conn.commit()
    pool.putconn(conn)

    return redirect("/items")
    
@items_bp.route("/api/items")
def api_items():
    item_type = str(request.args.get("type") or "all").strip().lower()
    category = str(request.args.get("category") or "all").strip()
    conn = get_db()
    
    cur = conn.cursor()

    where = ["items.company_id = %s"]
    params = [session.get("company_id")]
    if item_type in ("product", "service"):
        where.append("COALESCE(items.item_type, 'product') = %s")
        params.append(item_type)
    if category and category.lower() != "all":
        where.append("LOWER(COALESCE(items.category, '')) = LOWER(%s)")
        params.append(category)

    cur.execute(f"""
    SELECT 
        items.*,
        (SELECT image FROM item_images 
         WHERE item_id = items.id 
         LIMIT 1) as image
    FROM items
    WHERE {' AND '.join(where)}
    ORDER BY items.id DESC
    """, tuple(params))
    
    items = cur.fetchall()

    pool.putconn(conn)

    return jsonify([_catalog_item_payload(i) for i in items])


@items_bp.route("/api/categories")
def api_categories():
    category_type = str(request.args.get("type") or "all").strip().lower()
    conn = get_db()
    cur = conn.cursor()
    try:
        if category_type in ("product", "service"):
            cur.execute("""
                SELECT id, name, markup_percent, category_type
                FROM categories
                WHERE company_id = %s AND COALESCE(category_type, 'product') = %s
                ORDER BY name
            """, (session.get("company_id"), category_type))
        else:
            cur.execute("""
                SELECT id, name, markup_percent, category_type
                FROM categories WHERE company_id = %s ORDER BY category_type, name
            """, (session.get("company_id"),))
        return jsonify([_catalog_item_payload(row) for row in cur.fetchall()])
    finally:
        pool.putconn(conn)
    
@items_bp.route("/add_category", methods=["POST"])
def add_category():
    data = request.json

    conn = get_db()
    
    cur = conn.cursor()

    category_type = data.get("category_type", "product")
    if category_type not in ("product", "service"):
        category_type = "product"

    markup_percent = data.get("markup", 0) if category_type == "product" else 0

    cur.execute("""
        INSERT INTO categories
        (company_id, name, markup_percent, category_type)
        VALUES (%s, %s, %s, %s)
        RETURNING id, name, markup_percent, category_type
    """, (
        session.get("company_id"),
        data["name"],
        markup_percent,
        category_type
    ))

    category = cur.fetchone()

    conn.commit()
    pool.putconn(conn)

    return jsonify(dict(category))
    
@items_bp.route("/delete_category/<int:id>", methods=["POST"])
def delete_category(id):

    conn = get_db()
    
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM categories
        WHERE id = %s
        AND company_id = %s
    """, (
        id,
        session.get("company_id")
    ))

    conn.commit()

    pool.putconn(conn)

    return jsonify({"success": True})
    
@items_bp.route("/edit_category/<int:id>", methods=["POST"])
def edit_category(id):
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    category_type = data.get("category_type", "product")
    if category_type not in ("product", "service"):
        category_type = "product"

    markup_percent = data.get("markup", 0) if category_type == "product" else 0

    cur.execute("""
        UPDATE categories
        SET name = %s,
            markup_percent = %s,
            category_type = %s
        WHERE id = %s
          AND company_id = %s
        RETURNING id, name, markup_percent, category_type
    """, (
        data["name"],
        markup_percent,
        category_type,
        id,
        session.get("company_id")
    ))

    category = cur.fetchone()

    conn.commit()
    pool.putconn(conn)

    return jsonify(dict(category))
    
@items_bp.route("/api/barcode-info/<barcode>")
def barcode_info(barcode):

    import requests

    db = get_db()
    cur = db.cursor()

    item_type = request.args.get("item_type", "product")

    if item_type == "service":
        cur.execute("""
            SELECT *
            FROM items
            WHERE barcode = %s
              AND company_id = %s
              AND item_type = 'service'
        """, (
            barcode,
            session.get("company_id")
        ))
    else:
        cur.execute("""
            SELECT *
            FROM items
            WHERE barcode = %s
              AND company_id = %s
              AND COALESCE(item_type, 'product') = 'product'
        """, (
            barcode,
            session.get("company_id")
        ))

    item = cur.fetchone()

    # ✅ нашли локально
    if item:

        return jsonify({

            "found": True,

            "local": True,

            "name": item["name"],

            "category": item["category"],

            "price": item["retail_price"],
            
            "gtin": item.get("gtin"),
            
            "ntin": item.get("ntin"),
            
            "is_marked": item.get("is_marked"),
            "item_type": item.get("item_type") or "product",

        })

    if item_type == "service":
        pool.putconn(db)
        return jsonify({
            "found": False
        })

    pool.putconn(db)

    # 🌍 NATIONAL CATALOG
    try:

        url = (
            "https://e-catalog.gov.kz/"
            "api/integration/ofd/"
            f"search_ofd/?tin={barcode}"
        )

        TOKEN = os.getenv("NCT_API_TOKEN")
        
        print("TOKEN =", TOKEN[:30] if TOKEN else "NONE")

        headers = {
            "Authorization": f"JWT {TOKEN}"
        }

        response = requests.get(
            url,
            headers=headers,
            timeout=10
        )

        data = response.json()

        print("FULL DATA =", data)

        # 🔥 ЕСЛИ НАШЛИ
        if data:

            # 🔥 если results
            if "results" in data:

                if len(data["results"]) > 0:
                    product = data["results"][0]
                else:
                    product = {}

            # 🔥 если data
            elif "data" in data:

                product = data["data"]

            # 🔥 если список
            elif isinstance(data, list):

                product = data[0]

            # 🔥 обычный объект
            else:

                product = data

            return jsonify({

                "found": True,

                "external": True,

                "name":
                    product.get("name_ru", ""),
                    
                "measure":
                    product.get("measure", ""),

                "gtin":
                    product.get("gtin", ""),

                "ntin":
                    product.get("ntin_code", ""),
                    
                "is_marked":
                    product.get(
                        "is_markedeac",
                        False
                    ),

                "barcode":
                    barcode

            })

    except Exception as e:

        print("NCT ERROR:", e)

    return jsonify({
        "found": False
    })
    
@items_bp.route("/api/items/create", methods=["POST"])
def api_create_item():
    data = request.get_json(silent=True) or {}
    item_type = "service" if data.get("item_type") == "service" else "product"
    service_sale_mode = data.get("service_sale_mode") if item_type == "service" else None
    if service_sale_mode not in ("order", "booking", "request"):
        service_sale_mode = "order" if item_type == "service" else None

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO items (
            name,
            category,
            unit,
            type,
            description,
            retail_price,
            purchase_price,
            wholesale_price,
            discount_percent,
            barcode,
            gtin,
            ntin,
            is_marked,
            item_type,
            service_sale_mode,
            quantity,
            company_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        data.get("name", ""),
        data.get("category", ""),
        data.get("unit", "шт"),
        data.get("type", "piece"),
        data.get("description", ""),
        float(data.get("retail_price") or 0),
        float(data.get("purchase_price") or 0),
        float(data.get("wholesale_price") or 0),
        int(data.get("discount_percent") or 0),
        data.get("barcode", ""),
        data.get("gtin", ""),
        data.get("ntin", ""),
        bool(data.get("is_marked", False)) if item_type == "product" else False,
        item_type,
        service_sale_mode,
        float(data.get("quantity") or 0) if item_type == "product" else 0,
        session.get("company_id")
    ))

    item_id = cur.fetchone()["id"]

    # если количество больше 0 — создаём приход в движении товара
    quantity = float(data.get("quantity") or 0)
    purchase_price = float(data.get("purchase_price") or 0)

    if quantity > 0 and item_type == "product":
        cur.execute("""
            INSERT INTO stock_movements (
                company_id,
                item_id,
                movement_type,
                quantity,
                price,
                total,
                comment,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            session.get("company_id"),
            item_id,
            "income",
            quantity,
            purchase_price,
            quantity * purchase_price,
            "Первичный остаток при создании товара",
            datetime.now()
        ))

    conn.commit()

    cur.execute("""
        SELECT *
        FROM items
        WHERE id = %s
          AND company_id = %s
    """, (item_id, session.get("company_id")))

    item = cur.fetchone()
    pool.putconn(conn)

    return jsonify({
        "success": True,
        "id": item_id,
        "item": _catalog_item_payload(item)
    })


@items_bp.route("/api/items/<int:item_id>", methods=["PATCH"])
def api_update_item(item_id):
    data = request.get_json(silent=True) or {}
    item_type = "service" if data.get("item_type") == "service" else "product"
    service_sale_mode = data.get("service_sale_mode") if item_type == "service" else None
    if service_sale_mode not in ("order", "booking", "request"):
        service_sale_mode = "order" if item_type == "service" else None

    name = str(data.get("name") or "").strip()
    if not name:
        return jsonify({"success": False, "message": "Укажите название позиции"}), 400

    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE items SET
                name=%s, category=%s, unit=%s, type=%s, description=%s,
                retail_price=%s, purchase_price=%s, wholesale_price=%s,
                discount_percent=%s, barcode=%s, gtin=%s, ntin=%s,
                is_marked=%s, item_type=%s, service_sale_mode=%s
            WHERE id=%s AND company_id=%s
            RETURNING *
        """, (
            name,
            str(data.get("category") or "").strip(),
            str(data.get("unit") or ("услуга" if item_type == "service" else "шт")).strip(),
            str(data.get("type") or "piece").strip(),
            str(data.get("description") or "").strip(),
            float(data.get("retail_price") or 0),
            float(data.get("purchase_price") or 0),
            float(data.get("wholesale_price") or 0),
            int(float(data.get("discount_percent") or 0)),
            str(data.get("barcode") or "").strip(),
            str(data.get("gtin") or "").strip() if item_type == "product" else "",
            str(data.get("ntin") or "").strip() if item_type == "product" else "",
            bool(data.get("is_marked", False)) if item_type == "product" else False,
            item_type,
            service_sale_mode,
            item_id,
            session.get("company_id"),
        ))
        item = cur.fetchone()
        if not item:
            return jsonify({"success": False, "message": "Позиция не найдена"}), 404
        conn.commit()
        return jsonify({"success": True, "item": _catalog_item_payload(item)})
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@items_bp.route("/api/items/<int:item_id>", methods=["DELETE"])
def api_delete_item(item_id):
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM items WHERE id=%s AND company_id=%s RETURNING id",
            (item_id, session.get("company_id")),
        )
        deleted = cur.fetchone()
        if not deleted:
            return jsonify({"success": False, "message": "Позиция не найдена"}), 404
        conn.commit()
        return jsonify({"success": True})
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)

# ================== ИМПОРТ / ЭКСПОРТ КАТАЛОГА ==================

_CATALOG_HEADERS = [
    "Название", "Тип", "Категория", "Ед. изм.", "Штрихкод",
    "GTIN", "NTIN", "Закупочная цена", "Оптовая цена",
    "Розничная цена", "Скидка %", "Описание", "Количество"
]

_CATALOG_ALIASES = {
    "название": "name", "наименование": "name", "товар": "name",
    "тип": "item_type", "тип позиции": "item_type",
    "категория": "category",
    "ед. изм.": "unit", "ед изм": "unit", "единица измерения": "unit",
    "штрихкод": "barcode", "barcode": "barcode", "ean": "barcode",
    "gtin": "gtin", "ntin": "ntin",
    "закупочная цена": "purchase_price", "закуп": "purchase_price",
    "оптовая цена": "wholesale_price", "опт": "wholesale_price",
    "розничная цена": "retail_price", "цена": "retail_price",
    "скидка %": "discount_percent", "скидка": "discount_percent",
    "описание": "description", "количество": "quantity", "остаток": "quantity",
}


def _catalog_number(value, default=0):
    if value in (None, ""):
        return default
    try:
        normalized = str(value).strip().replace(" ", "").replace(",", ".")
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return default


def _catalog_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


_UNIT_ALIASES = {
    "штука": "шт", "штуки": "шт", "штук": "шт", "шт.": "шт", "piece": "шт", "pcs": "шт",
    "пара": "пар", "пары": "пар", "пар.": "пар", "pair": "пар",
    "комплект": "компл", "комплекта": "компл", "компл.": "компл", "к-т": "компл", "kit": "компл", "set": "компл",
    "упаковка": "упак", "упаковки": "упак", "упак.": "упак", "уп.": "упак", "pack": "упак",
    "пачка": "пач", "пач.": "пач", "коробка": "кор", "кор.": "кор", "короб": "кор", "box": "кор",
    "бутылка": "бут", "бут.": "бут", "канистра": "кан", "кан.": "кан", "рулон": "рул", "рул.": "рул", "roll": "рул",
    "килограмм": "кг", "килограммы": "кг", "кг.": "кг", "kg": "кг",
    "грамм": "г", "граммы": "г", "гр": "г", "гр.": "г", "g": "г", "gr": "г",
    "тонна": "т", "тонны": "т", "т.": "т", "ton": "т",
    "литр": "л", "литры": "л", "л.": "л", "liter": "л", "litre": "л", "l": "л",
    "миллилитр": "мл", "миллилитры": "мл", "мл.": "мл", "ml": "мл",
    "метр": "м", "метры": "м", "м.": "м", "пог.м": "м", "пог. м": "м", "погонный метр": "м", "meter": "м",
    "сантиметр": "см", "сантиметры": "см", "см.": "см", "cm": "см",
    "миллиметр": "мм", "миллиметры": "мм", "мм.": "мм", "mm": "мм",
    "квадратный метр": "м²", "кв.м": "м²", "кв. м": "м²", "м2": "м²", "m2": "м²", "m²": "м²",
    "кубический метр": "м³", "куб.м": "м³", "куб. м": "м³", "м3": "м³", "m3": "м³", "m³": "м³",
    "час": "час", "часа": "час", "часов": "час", "ч.": "час", "hour": "час",
    "день": "день", "дня": "день", "дней": "день", "сутки": "день", "day": "день",
    "неделя": "неделя", "недели": "неделя", "недель": "неделя", "нед.": "неделя", "week": "неделя",
    "месяц": "месяц", "месяца": "месяц", "месяцев": "месяц", "мес": "месяц", "мес.": "месяц", "month": "месяц",
    "год": "год", "года": "год", "лет": "год", "year": "год", "смены": "смена",
    "услуги": "услуга", "работа": "услуга", "service": "услуга",
    "чел": "человек", "чел.": "человек", "person": "человек", "места": "место", "мест": "место",
    "пассажира": "пассажир", "рейса": "рейс", "тура": "тур",
}


def _normalize_unit(value, default=""):
    unit = _catalog_text(value).lower().replace("ё", "е")
    unit = re.sub(r"\s+", " ", unit).strip()
    if not unit:
        return default
    # Неизвестное значение не заменяем на «шт»: сохраняем исходное для проверки.
    return _UNIT_ALIASES.get(unit, unit)


def _style_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="252B3A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    widths = [34, 14, 24, 13, 20, 18, 18, 18, 18, 18, 12, 40, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.auto_filter.ref = ws.dimensions


@items_bp.route("/items/export.xlsx")
def export_items_xlsx():
    company_id = session.get("company_id")
    if not company_id:
        return "Компания не выбрана", 403

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT name, item_type, category, unit, barcode, gtin, ntin,
                   purchase_price, wholesale_price, retail_price,
                   discount_percent, description, quantity
            FROM items
            WHERE company_id = %s
            ORDER BY COALESCE(item_type, 'product'), category, name
        """, (company_id,))
        rows = cur.fetchall()
    finally:
        pool.putconn(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    ws.append(_CATALOG_HEADERS)

    for row in rows:
        ws.append([
            row.get("name") or "",
            "Услуга" if (row.get("item_type") or "product") == "service" else "Товар",
            row.get("category") or "",
            row.get("unit") or "",
            row.get("barcode") or "",
            row.get("gtin") or "",
            row.get("ntin") or "",
            row.get("purchase_price") or 0,
            row.get("wholesale_price") or 0,
            row.get("retail_price") or 0,
            row.get("discount_percent") or 0,
            row.get("description") or "",
            row.get("quantity") or 0,
        ])

    _style_excel_sheet(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"nika_catalog_company_{company_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@items_bp.route("/items/import-template.xlsx")
def items_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    ws.append(_CATALOG_HEADERS)
    ws.append([
        "Масло моторное 5W-30", "Товар", "Масла", "шт",
        "4870000000000", "", "", 12000, 0, 15500, 0,
        "Пример товара. Эту строку можно удалить.", 10
    ])
    ws.append([
        "Замена масла", "Услуга", "Автосервис", "услуга",
        "", "", "", 0, 0, 5000, 0,
        "Пример услуги. Эту строку можно удалить.", 0
    ])
    _style_excel_sheet(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="nika_catalog_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@items_bp.route("/items/import", methods=["POST"])
def import_items_xlsx():
    company_id = session.get("company_id")
    if not company_id:
        return jsonify({"success": False, "message": "Компания не выбрана"}), 403

    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"success": False, "message": "Выберите Excel-файл"}), 400
    if not upload.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "message": "Поддерживается формат .xlsx"}), 400

    duplicate_mode = request.form.get("duplicate_mode", "skip")
    if duplicate_mode not in ("skip", "update"):
        duplicate_mode = "skip"

    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"success": False, "message": "Не удалось открыть Excel-файл"}), 400

    raw_headers = [_catalog_text(cell.value).lower() for cell in ws[1]]
    header_map = {}
    for column_index, header in enumerate(raw_headers, start=1):
        field = _CATALOG_ALIASES.get(header)
        if field:
            header_map[field] = column_index

    if "name" not in header_map:
        return jsonify({
            "success": False,
            "message": "В файле нет обязательной колонки «Название»"
        }), 400

    def value(row, field):
        column = header_map.get(field)
        return row[column - 1] if column else None

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    conn = get_db()
    try:
        cur = conn.cursor()
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = _catalog_text(value(row, "name"))
            if not name:
                continue

            try:
                raw_type = _catalog_text(value(row, "item_type")).lower()
                item_type = "service" if raw_type in ("услуга", "service", "работа") else "product"
                category = _catalog_text(value(row, "category")) or ("Услуги" if item_type == "service" else "Без категории")
                barcode = _catalog_text(value(row, "barcode"))
                gtin = _catalog_text(value(row, "gtin"))
                ntin = _catalog_text(value(row, "ntin"))
                unit = _normalize_unit(value(row, "unit"), "услуга" if item_type == "service" else "шт")
                purchase_price = _catalog_number(value(row, "purchase_price"))
                wholesale_price = _catalog_number(value(row, "wholesale_price"))
                retail_price = _catalog_number(value(row, "retail_price"))
                discount_percent = int(_catalog_number(value(row, "discount_percent")))
                description = _catalog_text(value(row, "description"))
                quantity = _catalog_number(value(row, "quantity")) if item_type == "product" else Decimal("0")

                cur.execute("""
                    INSERT INTO categories (company_id, name, markup_percent, category_type)
                    SELECT %s, %s, 0, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM categories
                        WHERE company_id = %s AND LOWER(name) = LOWER(%s)
                              AND category_type = %s
                    )
                """, (company_id, category, item_type, company_id, category, item_type))

                existing = None
                if barcode:
                    cur.execute("""
                        SELECT id FROM items
                        WHERE company_id = %s AND barcode = %s
                              AND COALESCE(item_type, 'product') = %s
                        LIMIT 1
                    """, (company_id, barcode, item_type))
                    existing = cur.fetchone()

                if not existing:
                    cur.execute("""
                        SELECT id FROM items
                        WHERE company_id = %s AND LOWER(name) = LOWER(%s)
                              AND COALESCE(item_type, 'product') = %s
                        LIMIT 1
                    """, (company_id, name, item_type))
                    existing = cur.fetchone()

                if existing and duplicate_mode == "skip":
                    stats["skipped"] += 1
                    continue

                if existing:
                    # При обновлении меняем только действительно заполненные
                    # колонки Excel. Пустые ячейки не должны стирать штрихкод,
                    # цены, остаток и остальные данные существующего товара.
                    updates = ["name=%s"]
                    update_values = [name]

                    text_fields = {
                        "category": category,
                        "unit": unit,
                        "description": description,
                        "barcode": barcode,
                        "gtin": gtin,
                        "ntin": ntin,
                    }
                    for field, field_value in text_fields.items():
                        if field in header_map and _catalog_text(value(row, field)):
                            updates.append(f"{field}=%s")
                            update_values.append(field_value)

                    numeric_fields = {
                        "retail_price": retail_price,
                        "wholesale_price": wholesale_price,
                        "purchase_price": purchase_price,
                        "discount_percent": discount_percent,
                    }
                    for field, field_value in numeric_fields.items():
                        raw_value = value(row, field)
                        if field in header_map and raw_value not in (None, ""):
                            updates.append(f"{field}=%s")
                            update_values.append(field_value)

                    if "quantity" in header_map and value(row, "quantity") not in (None, ""):
                        updates.append("quantity=%s")
                        update_values.append(quantity)

                    if "item_type" in header_map and raw_type:
                        updates.extend([
                            "item_type=%s",
                            "is_marked=CASE WHEN %s='service' THEN FALSE ELSE COALESCE(is_marked, FALSE) END",
                        ])
                        update_values.extend([item_type, item_type])

                    update_values.extend([existing["id"], company_id])
                    cur.execute(
                        f"UPDATE items SET {', '.join(updates)} "
                        "WHERE id=%s AND company_id=%s",
                        tuple(update_values),
                    )
                    stats["updated"] += 1
                else:
                    cur.execute("""
                        INSERT INTO items (
                            name, category, unit, description, retail_price,
                            wholesale_price, purchase_price, discount_percent,
                            barcode, gtin, ntin, is_marked, item_type,
                            service_sale_mode, quantity, company_id
                        ) VALUES (
                            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,FALSE,%s,%s,%s,%s
                        )
                    """, (
                        name, category, unit, description, retail_price,
                        wholesale_price, purchase_price, discount_percent,
                        barcode, gtin, ntin, item_type,
                        "order" if item_type == "service" else None,
                        quantity, company_id
                    ))
                    stats["created"] += 1
            except Exception as row_error:
                stats["errors"].append({"row": excel_row, "message": str(row_error)[:180]})
                if len(stats["errors"]) >= 50:
                    break

        conn.commit()
        return jsonify({"success": True, **stats})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "message": str(error)}), 500
    finally:
        pool.putconn(conn)


# ================== ОБОГАЩЕНИЕ КАТАЛОГА НКТ ==================

def _nct_normalize_name(value):
    value = (value or "").lower().replace("ё", "е")
    value = re.sub(r"[^0-9a-zа-яәіңғүұқөһ]+", " ", value, flags=re.IGNORECASE)
    return " ".join(value.split())


def _nct_search_queries(value):
    """Формирует несколько вариантов запроса от точного к более широкому."""
    original = _catalog_text(value)
    if not original:
        return []

    # Убираем внутренний код 1С в начале: (001NB), (002VP) и т.п.
    cleaned = re.sub(r"^\s*\([^)]{1,30}\)\s*", "", original).strip()

    variants = [cleaned]

    # Нормализуем похожие кириллические буквы внутри артикулов.
    # Например 281133М000 -> 281133M000.
    latinized = cleaned.translate(str.maketrans({
        "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M",
        "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
        "а": "a", "е": "e", "к": "k", "м": "m", "о": "o",
        "р": "p", "с": "c", "т": "t", "х": "x",
    }))
    variants.append(latinized)

    # Артикулы и модельные коды часто дают самый точный результат.
    article_tokens = re.findall(
        r"(?i)\b(?=[A-ZА-Я0-9-]*\d)[A-ZА-Я0-9]+(?:[-/][A-ZА-Я0-9]+)*\b",
        cleaned
    )
    variants.extend(reversed(article_tokens))

    # Полезные укороченные варианты без общих служебных слов.
    words = cleaned.split()
    if len(words) > 5:
        variants.append(" ".join(words[:5]))
    if len(words) > 3:
        variants.append(" ".join(words[:3]))

    stop_words = {
        "передний", "задний", "левый", "правый", "без", "для",
        "в", "на", "и", "комплект", "шт", "новый"
    }
    meaningful = [word for word in words if word.lower() not in stop_words]
    if meaningful:
        variants.append(" ".join(meaningful))

    unique = []
    seen = set()
    for variant in variants:
        variant = re.sub(r"\s+", " ", variant).strip(" ,.;:-")
        key = variant.casefold()
        if len(variant) >= 2 and key not in seen:
            seen.add(key)
            unique.append(variant)

    return unique[:8]


def _nct_extract_list(payload):
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        return []
    # НКТ возвращает список товаров в поле "result" (в единственном числе).
    # Поддерживаем также другие возможные варианты оболочки ответа.
    for key in ("result", "results", "data", "items", "products", "content"):
        value = payload.get(key)
        if isinstance(value, list):
            return value
        if isinstance(value, dict):
            nested = _nct_extract_list(value)
            if nested:
                return nested
    return [payload] if payload else []


def _nct_value(product, *keys):
    for key in keys:
        value = product.get(key) if isinstance(product, dict) else None
        if value not in (None, ""):
            return value
    return ""


def _nct_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "да"}


def _nct_search_by_name(query):
    """Ищет товар по нескольким вариантам наименования в НКТ."""
    search_url = "https://nct.gov.kz/api/integration/ofd/search_ofd/"
    token = (os.getenv("NCT_API_TOKEN") or "").strip()
    if not token:
        return {
            "configured": False,
            "message": "Не задан NCT_API_TOKEN. Добавьте действующий ключ НКТ в .env и полностью перезапустите Nika."
        }

    all_products = []
    used_queries = []
    last_status = None

    for search_query in _nct_search_queries(query):
        response = None

        for auth_scheme in ("JWT", "Bearer"):
            response = requests.get(
                search_url,
                params={"q": search_query},
                headers={
                    "Accept": "application/json",
                    "Authorization": f"{auth_scheme} {token}",
                },
                timeout=20,
            )
            if response.status_code != 401:
                break

        if response is None:
            continue

        last_status = response.status_code
        if response.status_code >= 400:
            details = (response.text or "").strip().replace("\n", " ")
            if len(details) > 300:
                details = details[:300] + "..."
            raise RuntimeError(
                f"НКТ вернул HTTP {response.status_code}"
                + (f": {details}" if details else "")
            )

        try:
            payload = response.json()
        except ValueError as exc:
            raise RuntimeError("НКТ вернул ответ не в формате JSON") from exc

        products = _nct_extract_list(payload)
        used_queries.append({
            "query": search_query,
            "found": len(products),
        })
        all_products.extend(products)

        # Если точный или артикульный запрос уже дал достаточно результатов,
        # не перегружаем API лишними обращениями.
        if len(all_products) >= 20:
            break

    source_name = _nct_normalize_name(
        re.sub(r"^\s*\([^)]{1,30}\)\s*", "", query or "")
    )
    candidates = []

    for product in all_products:
        if not isinstance(product, dict):
            continue
        if _nct_bool(_nct_value(product, "ntin_isdeactivated", "is_deactivated")):
            continue

        name = _nct_value(
            product,
            "name_ru", "name", "product_name", "title", "full_name",
            "trade_name", "short_name"
        )
        if not name:
            continue

        candidate_name = _nct_normalize_name(str(name))
        score = round(SequenceMatcher(None, source_name, candidate_name).ratio() * 100)

        # Совпадение артикула сильно важнее полного текстового сходства.
        source_articles = set(re.findall(r"\b[a-zа-я0-9]+(?:[-/][a-zа-я0-9]+)*\b", source_name))
        candidate_articles = set(re.findall(r"\b[a-zа-я0-9]+(?:[-/][a-zа-я0-9]+)*\b", candidate_name))
        shared_articles = {
            token for token in source_articles & candidate_articles
            if any(ch.isdigit() for ch in token) and len(token) >= 4
        }
        if shared_articles:
            score = max(score, 95)

        gtin = _catalog_text(_nct_value(product, "gtin", "GTIN", "barcode", "ean"))
        ntin = _catalog_text(_nct_value(product, "ntin_code", "ntin", "NTIN", "kztin"))

        candidates.append({
            "name": _catalog_text(name),
            "name_kk": _catalog_text(_nct_value(product, "name_kk")),
            "gtin": gtin,
            "ntin": ntin,
            "barcode": gtin or _catalog_text(_nct_value(product, "barcode", "ean")),
            "measure": _catalog_text(_nct_value(product, "measure", "unit", "measure_name")),
            "is_marked": _nct_bool(_nct_value(product, "is_markedeac", "is_marked", "marked")),
            "is_social": _nct_bool(_nct_value(product, "is_social")),
            "modified": _catalog_text(_nct_value(product, "modified")),
            "manufacturer": _catalog_text(
                _nct_value(product, "manufacturer_name", "manufacturer", "producer")
            ),
            "score": score,
        })

    candidates.sort(key=lambda item: item["score"], reverse=True)
    unique = []
    seen = set()

    for candidate in candidates:
        identity = (
            candidate.get("gtin"),
            candidate.get("ntin"),
            candidate.get("name"),
        )
        if identity in seen:
            continue
        seen.add(identity)
        unique.append(candidate)
        if len(unique) >= 10:
            break

    print("NCT SEARCH DEBUG:", {
        "source": query,
        "queries": used_queries,
        "status": last_status,
        "candidates": len(unique),
    })

    return {
        "configured": True,
        "candidates": unique,
        "queries": used_queries,
    }


@items_bp.route("/api/catalog/enrichment/items")
def catalog_enrichment_items():
    company_id = session.get("company_id")
    limit = min(max(int(request.args.get("limit", 25)), 1), 100)
    only_missing = request.args.get("only_missing", "1") != "0"

    conn = get_db()
    try:
        cur = conn.cursor()
        missing_sql = "AND (COALESCE(gtin, '') = '' OR COALESCE(ntin, '') = '')" if only_missing else ""
        cur.execute(f"""
            SELECT id, name, barcode, gtin, ntin, unit
            FROM items
            WHERE company_id = %s
              AND COALESCE(item_type, 'product') = 'product'
              {missing_sql}
            ORDER BY id
            LIMIT %s
        """, (company_id, limit))
        rows = [dict(row) for row in cur.fetchall()]
        return jsonify({"success": True, "items": rows, "count": len(rows)})
    finally:
        pool.putconn(conn)


@items_bp.route("/api/catalog/enrichment/search", methods=["POST"])
def catalog_enrichment_search():
    data = request.get_json(silent=True) or {}
    item_id = data.get("item_id")
    company_id = session.get("company_id")

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, name, barcode, gtin, ntin, unit
            FROM items
            WHERE id = %s AND company_id = %s
              AND COALESCE(item_type, 'product') = 'product'
        """, (item_id, company_id))
        item = cur.fetchone()
    finally:
        pool.putconn(conn)

    if not item:
        return jsonify({"success": False, "message": "Товар не найден"}), 404

    try:
        result = _nct_search_by_name(item["name"])
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 502

    if not result.get("configured"):
        return jsonify({"success": False, "configured": False, "message": result["message"]}), 503

    return jsonify({
        "success": True,
        "item": dict(item),
        "candidates": result.get("candidates", [])
    })


@items_bp.route("/api/catalog/enrichment/apply", methods=["POST"])
def catalog_enrichment_apply():
    data = request.get_json(silent=True) or {}
    item_id = data.get("item_id")
    candidate = data.get("candidate") or {}
    fields = data.get("fields") or {}
    company_id = session.get("company_id")

    allowed = {
        "gtin": _catalog_text(candidate.get("gtin")),
        "ntin": _catalog_text(candidate.get("ntin")),
        "barcode": _catalog_text(candidate.get("barcode")),
        "unit": _catalog_text(candidate.get("measure")),
        "name": _catalog_text(candidate.get("name")),
        "is_marked": bool(candidate.get("is_marked")),
    }
    updates, values = [], []
    for field in ("gtin", "ntin", "barcode", "unit", "name", "is_marked"):
        if fields.get(field) and allowed[field] not in (None, ""):
            updates.append(f"{field} = %s")
            values.append(allowed[field])

    if not updates:
        return jsonify({"success": False, "message": "Не выбраны данные для обновления"}), 400

    conn = get_db()
    try:
        cur = conn.cursor()
        values.extend([item_id, company_id])
        cur.execute(f"""
            UPDATE items
            SET {', '.join(updates)}
            WHERE id = %s AND company_id = %s
              AND COALESCE(item_type, 'product') = 'product'
            RETURNING id, name, barcode, gtin, ntin, unit, is_marked
        """, values)
        updated = cur.fetchone()
        if not updated:
            conn.rollback()
            return jsonify({"success": False, "message": "Товар не найден"}), 404
        conn.commit()
        return jsonify({"success": True, "item": dict(updated)})
    except Exception as exc:
        conn.rollback()
        return jsonify({"success": False, "message": str(exc)}), 500
    finally:
        pool.putconn(conn)
