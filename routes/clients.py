from datetime import datetime, timedelta
from io import BytesIO
import json
import os
import re
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from flask import (
    Blueprint,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
    send_file,
)
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import get_db, pool
from utils.timezone import now_kz


UPLOAD_DIR = os.path.join("static", "uploads", "clients")
COMMENT_UPLOAD_DIR = os.path.join(UPLOAD_DIR, "comments")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(COMMENT_UPLOAD_DIR, exist_ok=True)

clients_bp = Blueprint("clients", __name__)


def format_date_ru(date_value):
    if not date_value:
        return ""

    if hasattr(date_value, "strftime"):
        return date_value.strftime("%d.%m.%Y")

    try:
        return datetime.strptime(str(date_value), "%Y-%m-%d").strftime("%d.%m.%Y")
    except (TypeError, ValueError):
        return str(date_value)


def _company_id():
    return session.get("company_id")


def _save_uploaded_file(file_storage, folder, with_microseconds=False):
    if not file_storage or not file_storage.filename:
        return ""

    safe_name = secure_filename(file_storage.filename)
    if not safe_name:
        return ""

    date_format = "%Y%m%d%H%M%S%f" if with_microseconds else "%Y%m%d%H%M%S"
    filename = f"{now_kz().strftime(date_format)}_{safe_name}"
    save_path = os.path.join(folder, filename)
    file_storage.save(save_path)

    return "/" + save_path.replace("\\", "/")


def _serialize_value(value):
    if isinstance(value, (datetime,)):
        return value.isoformat()
    return value


def _serialize_row(row):
    if not row:
        return {}
    return {key: _serialize_value(value) for key, value in dict(row).items()}


_IDENTIFIER_WEIGHTS_PRIMARY = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11)
_IDENTIFIER_WEIGHTS_SECONDARY = (3, 4, 5, 6, 7, 8, 9, 10, 11, 1, 2)


def _normalize_identifier(value):
    return re.sub(r"\D", "", str(value or ""))


def _valid_kz_identifier(value):
    """Validate the checksum shared by Kazakhstan IIN and BIN values."""
    identifier = _normalize_identifier(value)
    if len(identifier) != 12:
        return False

    digits = [int(char) for char in identifier]
    checksum = sum(
        digit * weight
        for digit, weight in zip(digits[:11], _IDENTIFIER_WEIGHTS_PRIMARY)
    ) % 11
    if checksum == 10:
        checksum = sum(
            digit * weight
            for digit, weight in zip(digits[:11], _IDENTIFIER_WEIGHTS_SECONDARY)
        ) % 11
    return checksum != 10 and checksum == digits[11]


def _first_value(data, *keys):
    if not isinstance(data, dict):
        return ""

    normalized = {
        re.sub(r"[^a-zа-я0-9]", "", str(key).lower()): value
        for key, value in data.items()
    }
    for key in keys:
        value = normalized.get(re.sub(r"[^a-zа-я0-9]", "", key.lower()))
        if value not in (None, "", [], {}):
            return str(value).strip()
    return ""


def _person_name(value):
    """Return a readable name from KGD fullName values (object or string)."""
    if isinstance(value, str):
        return value.strip()
    if not isinstance(value, dict):
        return ""

    parts = []
    for keys in (
        ("lastName", "last_name", "surname"),
        ("firstName", "first_name", "name"),
        ("middleName", "middle_name", "patronymic"),
    ):
        part = _first_value(value, *keys)
        if part:
            parts.append(part)
    return " ".join(parts)


def _unwrap_lookup_records(payload):
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        return []

    for key in (
        "taxpayerPortalSearchResponses",
        "data", "result", "results", "items", "records", "taxpayer",
    ):
        nested = payload.get(key)
        if isinstance(nested, list):
            return nested
        if isinstance(nested, dict):
            return [nested]
    return [payload]


def _map_lookup_record(record, identifier):
    if not isinstance(record, dict):
        return None
    if isinstance(record.get("_source"), dict):
        record = record["_source"]

    record_identifier = _normalize_identifier(
        _first_value(
            record, "bin", "iin", "iinbin", "identifier", "taxpayerCode", "code"
        )
    )
    if record_identifier and record_identifier != identifier:
        return None

    company_name = _first_value(
        record,
        "nameru", "name_ru", "company_name", "companyName", "organizationName",
        "taxpayerName", "name", "namekz", "name_kz",
    )
    raw_full_name = record.get("fullName")
    full_name = _person_name(raw_full_name) or _first_value(
        record,
        "headru", "head_name", "headName", "leaderFio", "directorFio",
        "director", "fio", "full_name",
    )
    address = _first_value(
        record,
        "addressru", "address_ru", "legal_address", "legalAddress", "address",
        "location", "addresskz", "address_kz",
    )
    phone = _first_value(record, "phone", "telephone", "mobile", "phoneNumber")

    taxpayer_type = _first_value(record, "taxpayerType", "taxpayer_type")
    if not company_name and taxpayer_type in ("IP", "UL", "UL_NR"):
        company_name = full_name

    if not any((company_name, full_name, address, phone)):
        return None

    return {
        "iin": identifier,
        "company_name": company_name,
        "full_name": full_name or company_name,
        "address": address,
        "phone": phone,
        "taxpayer_type": taxpayer_type,
        "registration_begin_date": _first_value(record, "beginDate", "begin_date"),
        "registration_end_date": _first_value(record, "endDate", "end_date"),
    }


def _fetch_json(url, headers=None, timeout=8):
    request_object = Request(
        url,
        headers={"Accept": "application/json", "User-Agent": "NikaBusiness/1.0", **(headers or {})},
    )
    with urlopen(request_object, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def _lookup_configured_provider(identifier):
    url_template = os.getenv("CLIENT_LOOKUP_API_URL", "").strip()
    if not url_template:
        return None

    separator = "&" if "?" in url_template else "?"
    if "{identifier}" in url_template:
        url = url_template.replace("{identifier}", identifier)
    else:
        url = f"{url_template}{separator}{urlencode({'identifier': identifier})}"

    headers = {}
    token = os.getenv("CLIENT_LOOKUP_API_TOKEN", "").strip()
    if token:
        header_name = os.getenv("CLIENT_LOOKUP_API_TOKEN_HEADER", "X-Portal-Token").strip()
        headers[header_name or "X-Portal-Token"] = token

    payload = _fetch_json(url, headers=headers)
    for record in _unwrap_lookup_records(payload):
        mapped = _map_lookup_record(record, identifier)
        if mapped:
            return mapped
    return None


def _lookup_kgd_taxpayer(identifier):
    """Look up a taxpayer using the official KGD ISNA Portal API."""
    token = os.getenv("KGD_PORTAL_TOKEN", "").strip()
    if not token:
        return None

    base_url = os.getenv(
        "KGD_TAXPAYER_API_URL",
        "https://portal.kgd.gov.kz/services/isnaportalsync/public/taxpayer-data",
    ).strip()
    if not base_url:
        return None

    # One 12-digit identifier can belong to a legal entity, an individual
    # entrepreneur, or a natural person. KGD requires the type explicitly.
    taxpayer_types = ("UL", "IP", "FL")
    headers = {"X-Portal-Token": token}
    timeout = max(1, min(int(os.getenv("KGD_API_TIMEOUT", "8")), 30))

    for taxpayer_type in taxpayer_types:
        separator = "&" if "?" in base_url else "?"
        url = base_url + separator + urlencode({
            "taxpayerCode": identifier,
            "taxpayerType": taxpayer_type,
            "print": "false",
        })
        payload = _fetch_json(url, headers=headers, timeout=timeout)
        for record in _unwrap_lookup_records(payload):
            if str(record.get("messageResult", "SUCCESS")).upper() != "SUCCESS":
                continue
            mapped = _map_lookup_record(record, identifier)
            if mapped:
                return mapped
    return None


def _lookup_egov_open_data(identifier):
    api_key = os.getenv("EGOV_OPEN_DATA_API_KEY", "").strip()
    if not api_key:
        return None

    source = {
        "size": 5,
        "query": {"bool": {"must": [{"match": {"bin": identifier}}]}},
    }
    url = "https://data.egov.kz/api/v4/gbd_ul/v1?" + urlencode(
        {"apiKey": api_key, "source": json.dumps(source, separators=(",", ":"))}
    )
    payload = _fetch_json(url)
    for record in _unwrap_lookup_records(payload):
        mapped = _map_lookup_record(record, identifier)
        if mapped:
            return mapped
    return None


def _merge_lookup_data(primary, secondary):
    """Fill empty registry fields without overwriting the primary source."""
    result = dict(primary or {})
    for key, value in (secondary or {}).items():
        if value not in (None, "", [], {}) and result.get(key) in (None, "", [], {}):
            result[key] = value
    return result


@clients_bp.route("/clients")
def clients():
    company_id = _company_id()
    search = request.args.get("search", "").strip()

    conn = get_db()
    try:
        cur = conn.cursor()

        search_condition = ""
        params = [company_id]

        if search:
            search_condition = """
                AND (
                    COALESCE(full_name, '') ILIKE %s
                    OR COALESCE(phone, '') ILIKE %s
                    OR COALESCE(company_name, '') ILIKE %s
                    OR COALESCE(iin, '') ILIKE %s
                )
            """
            pattern = f"%{search}%"
            params.extend([pattern, pattern, pattern, pattern])

        cur.execute(
            f"""
            SELECT *
            FROM clients
            WHERE company_id = %s
              AND COALESCE(is_deleted, FALSE) = FALSE
              {search_condition}
            ORDER BY id DESC
            """,
            tuple(params),
        )
        active_clients = cur.fetchall()

        deleted_params = [company_id]
        if search:
            deleted_params.extend([pattern, pattern, pattern, pattern])

        cur.execute(
            f"""
            SELECT *
            FROM clients
            WHERE company_id = %s
              AND COALESCE(is_deleted, FALSE) = TRUE
              {search_condition}
            ORDER BY id DESC
            """,
            tuple(deleted_params),
        )
        deleted_clients = cur.fetchall()

        return render_template(
            "clients.html",
            clients=active_clients,
            deleted_clients=deleted_clients,
            search=search,
        )
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/add", methods=["POST"])
def add_client():
    company_id = _company_id()

    full_name = request.form.get("full_name", "").strip()
    if not full_name:
        return jsonify({"status": "error", "message": "Укажите имя клиента"}), 400

    photo_path = _save_uploaded_file(request.files.get("photo"), UPLOAD_DIR)

    comment_photo_paths = []
    for file_storage in request.files.getlist("comment_photos"):
        path = _save_uploaded_file(
            file_storage,
            COMMENT_UPLOAD_DIR,
            with_microseconds=True,
        )
        if path:
            comment_photo_paths.append(path)

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO clients (
                full_name,
                phone,
                iin,
                company_name,
                status,
                category,
                payment,
                comment,
                address,
                contract_number,
                contract_date,
                photo,
                comment_photos,
                created_at,
                company_id,
                is_deleted
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, NULLIF(%s, '')::date, %s, %s, %s, %s, FALSE
            )
            RETURNING id
            """,
            (
                full_name,
                request.form.get("phone", "").strip(),
                request.form.get("iin", "").strip(),
                request.form.get("company_name", "").strip(),
                request.form.get("status", "Новый").strip() or "Новый",
                request.form.get("category", "").strip(),
                request.form.get("payment", "Не оплачено").strip() or "Не оплачено",
                request.form.get("comment", "").strip(),
                request.form.get("address", "").strip(),
                request.form.get("contract_number", "").strip(),
                request.form.get("contract_date", "").strip(),
                photo_path,
                "|".join(comment_photo_paths),
                now_kz(),
                company_id,
            ),
        )
        client_id = cur.fetchone()["id"]
        conn.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"status": "ok", "id": client_id})

        return redirect(url_for("clients.clients"))
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/<int:client_id>")
def client_detail(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM clients
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        client = cur.fetchone()

        if not client:
            return "Клиент не найден", 404

        cur.execute(
            """
            SELECT *
            FROM sales
            WHERE client_id = %s AND company_id = %s
            ORDER BY id DESC
            """,
            (client_id, company_id),
        )
        sales = cur.fetchall()

        formatted_sales = []
        for sale in sales:
            new_sale = dict(sale)
            created_at = sale.get("created_at") if hasattr(sale, "get") else sale["created_at"]

            try:
                new_sale["date_str"] = (created_at + timedelta(hours=5)).strftime(
                    "%d.%m.%Y %H:%M"
                )
            except (TypeError, AttributeError):
                new_sale["date_str"] = str(created_at or "")

            formatted_sales.append(new_sale)

        return render_template(
            "client_detail.html",
            client=client,
            sales=formatted_sales,
        )
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/<int:client_id>/add_item", methods=["POST"])
def add_item(client_id):
    company_id = _company_id()
    item_id = request.form.get("item_id", type=int)
    payment_method = request.form.get("payment_method", "Не оплачено")

    if not item_id:
        return jsonify({"status": "error", "message": "Товар не выбран"}), 400

    conn = get_db()
    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT id
            FROM clients
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        if not cur.fetchone():
            return jsonify({"status": "error", "message": "Клиент не найден"}), 404

        cur.execute(
            """
            SELECT *
            FROM items
            WHERE id = %s AND company_id = %s
            """,
            (item_id, company_id),
        )
        item = cur.fetchone()

        if not item:
            return jsonify({"status": "error", "message": "Товар не найден"}), 404

        cur.execute(
            """
            INSERT INTO client_items (
                client_id,
                item_id,
                price,
                payment_method,
                is_paid,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                client_id,
                item_id,
                item["price"],
                payment_method,
                0,
                now_kz(),
            ),
        )
        conn.commit()

        return jsonify({"status": "ok"})
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/client/<int:client_id>")
def api_client(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM clients
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        client = cur.fetchone()

        if not client:
            return jsonify({"status": "error", "message": "Клиент не найден"}), 404

        cur.execute(
            """
            SELECT *
            FROM sales
            WHERE client_id = %s AND company_id = %s
            ORDER BY id DESC
            """,
            (client_id, company_id),
        )
        deals = cur.fetchall()

        cur.execute(
            """
            SELECT *
            FROM items
            WHERE company_id = %s
            ORDER BY id DESC
            """,
            (company_id,),
        )
        items = cur.fetchall()

        return jsonify(
            {
                "client": _serialize_row(client),
                "deals": [_serialize_row(row) for row in deals],
                "items": [_serialize_row(row) for row in items],
            }
        )
    finally:
        pool.putconn(conn)




@clients_bp.route("/api/client/<int:client_id>/update", methods=["PATCH", "POST"])
def api_update_client(client_id):
    """Partially update a client without reloading the clients page."""
    company_id = _company_id()
    data = request.get_json(silent=True) or {}

    allowed_fields = {
        "full_name", "phone", "iin", "company_name", "status", "category",
        "payment", "comment", "address", "contract_number", "contract_date",
    }
    updates = {key: data.get(key) for key in allowed_fields if key in data}

    if not updates:
        return jsonify({"status": "error", "message": "Нет данных для сохранения"}), 400

    if "full_name" in updates and not str(updates["full_name"] or "").strip():
        return jsonify({"status": "error", "message": "Укажите имя клиента"}), 400

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT id FROM clients WHERE id = %s AND company_id = %s",
            (client_id, company_id),
        )
        if not cur.fetchone():
            return jsonify({"status": "error", "message": "Клиент не найден"}), 404

        set_parts = []
        params = []
        for field, value in updates.items():
            if field == "contract_date":
                set_parts.append("contract_date = NULLIF(%s, '')::date")
                params.append(str(value or "").strip())
            else:
                set_parts.append(f"{field} = %s")
                params.append(str(value or "").strip())

        params.extend([client_id, company_id])
        cur.execute(
            f"UPDATE clients SET {', '.join(set_parts)} WHERE id = %s AND company_id = %s RETURNING *",
            tuple(params),
        )
        updated = cur.fetchone()
        conn.commit()
        return jsonify({"status": "ok", "client": _serialize_row(updated)})
    except Exception as error:
        conn.rollback()
        return jsonify({"status": "error", "message": str(error)}), 500
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/<int:client_id>/edit", methods=["POST"])
def edit_client(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()

        cur.execute(
            """
            SELECT *
            FROM clients
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        old_client = cur.fetchone()

        if not old_client:
            return jsonify({"status": "error", "message": "Клиент не найден"}), 404

        full_name = request.form.get("full_name", "").strip()
        if not full_name:
            return jsonify({"status": "error", "message": "Укажите имя клиента"}), 400

        photo_path = old_client["photo"] or ""
        new_photo_path = _save_uploaded_file(request.files.get("photo"), UPLOAD_DIR)
        if new_photo_path:
            photo_path = new_photo_path

        old_comment_photos = old_client["comment_photos"] or ""
        comment_photo_paths = [
            path for path in old_comment_photos.split("|") if path
        ]

        for file_storage in request.files.getlist("comment_photos"):
            path = _save_uploaded_file(
                file_storage,
                COMMENT_UPLOAD_DIR,
                with_microseconds=True,
            )
            if path:
                comment_photo_paths.append(path)

        cur.execute(
            """
            UPDATE clients
            SET full_name = %s,
                phone = %s,
                iin = %s,
                company_name = %s,
                status = %s,
                category = %s,
                payment = %s,
                comment = %s,
                address = %s,
                contract_number = %s,
                contract_date = NULLIF(%s, '')::date,
                photo = %s,
                comment_photos = %s
            WHERE id = %s AND company_id = %s
            """,
            (
                full_name,
                request.form.get("phone", "").strip(),
                request.form.get("iin", "").strip(),
                request.form.get("company_name", "").strip(),
                request.form.get("status", "Новый").strip() or "Новый",
                request.form.get("category", "").strip(),
                request.form.get("payment", "Не оплачено").strip() or "Не оплачено",
                request.form.get("comment", "").strip(),
                request.form.get("address", "").strip(),
                request.form.get("contract_number", "").strip(),
                request.form.get("contract_date", "").strip(),
                photo_path,
                "|".join(comment_photo_paths),
                client_id,
                company_id,
            ),
        )
        conn.commit()

        return jsonify({"status": "ok"})
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/<int:client_id>/delete", methods=["POST", "GET"])
def delete_client(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE clients
            SET is_deleted = TRUE
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        conn.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"status": "ok"})

        return redirect(url_for("clients.clients"))
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/clients/deleted")
def deleted_clients():
    return redirect(url_for("clients.clients", tab="deleted"))


@clients_bp.route("/clients/<int:client_id>/restore", methods=["POST", "GET"])
def restore_client(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE clients
            SET is_deleted = FALSE
            WHERE id = %s AND company_id = %s
            """,
            (client_id, company_id),
        )
        conn.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"status": "ok"})

        return redirect(url_for("clients.clients", tab="deleted"))
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route(
    "/clients/<int:client_id>/delete_permanently",
    methods=["POST", "GET"],
)
def delete_client_permanently(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            DELETE FROM clients
            WHERE id = %s
              AND company_id = %s
              AND COALESCE(is_deleted, FALSE) = TRUE
            """,
            (client_id, company_id),
        )
        conn.commit()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"status": "ok"})

        return redirect(url_for("clients.clients", tab="deleted"))
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/clients")
def api_clients():
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                id,
                full_name,
                phone,
                iin,
                company_name,
                address
            FROM clients
            WHERE COALESCE(is_deleted, FALSE) = FALSE
              AND company_id = %s
            ORDER BY id DESC
            """,
            (company_id,),
        )
        rows = cur.fetchall()
        return jsonify([_serialize_row(row) for row in rows])
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/client/<int:client_id>/sales")
def client_sales(client_id):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT *
            FROM sales
            WHERE client_id = %s AND company_id = %s
            ORDER BY id DESC
            """,
            (client_id, company_id),
        )
        rows = cur.fetchall()
        return jsonify([_serialize_row(row) for row in rows])
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/clients/create", methods=["POST"])
def api_create_client():
    data = request.get_json(silent=True) or {}
    company_id = _company_id()

    full_name = str(data.get("full_name", "")).strip()
    if not full_name:
        return jsonify({"success": False, "message": "Укажите имя клиента"}), 400

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO clients (
                full_name,
                phone,
                iin,
                company_name,
                address,
                comment,
                status,
                payment,
                created_at,
                company_id,
                is_deleted
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, FALSE)
            RETURNING id
            """,
            (
                full_name,
                str(data.get("phone", "")).strip(),
                str(data.get("iin", "")).strip(),
                str(data.get("company_name", "")).strip(),
                str(data.get("address", "")).strip(),
                str(data.get("comment", "")).strip(),
                str(data.get("status", "Новый")).strip() or "Новый",
                str(data.get("payment", "Не оплачено")).strip() or "Не оплачено",
                now_kz(),
                company_id,
            ),
        )
        client_id = cur.fetchone()["id"]
        conn.commit()

        return jsonify({"success": True, "id": client_id})
    except Exception:
        conn.rollback()
        raise
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/clients/by-iin/<iin>")
def get_client_by_iin(iin):
    company_id = _company_id()
    conn = get_db()

    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT *
            FROM clients
            WHERE company_id = %s
              AND iin = %s
              AND COALESCE(is_deleted, FALSE) = FALSE
            LIMIT 1
            """,
            (company_id, iin),
        )
        client = cur.fetchone()

        if client:
            return jsonify({"found": True, "client": _serialize_row(client)})

        return jsonify({"found": False})
    finally:
        pool.putconn(conn)


@clients_bp.route("/api/clients/lookup", methods=["POST"])
def lookup_client_identifier():
    company_id = _company_id()
    if not company_id:
        return jsonify({"found": False, "message": "Компания не выбрана"}), 403

    payload = request.get_json(silent=True) or {}
    identifier = _normalize_identifier(payload.get("identifier"))
    if not _valid_kz_identifier(identifier):
        return jsonify({
            "found": False,
            "message": "Проверьте ИИН/БИН: нужен корректный номер из 12 цифр",
        }), 400

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT *
            FROM clients
            WHERE company_id = %s
              AND REGEXP_REPLACE(COALESCE(iin, ''), '\\D', '', 'g') = %s
              AND COALESCE(is_deleted, FALSE) = FALSE
            ORDER BY id DESC
            LIMIT 1
            """,
            (company_id, identifier),
        )
        client = cur.fetchone()
        if client:
            return jsonify({
                "found": True,
                "source": "local",
                "client_id": client["id"],
                "data": _serialize_row(client),
                "message": "Клиент уже есть в базе — данные подставлены",
            })
    finally:
        pool.putconn(conn)

    provider_configured = bool(
        os.getenv("KGD_PORTAL_TOKEN", "").strip()
        or
        os.getenv("CLIENT_LOOKUP_API_URL", "").strip()
        or os.getenv("EGOV_OPEN_DATA_API_KEY", "").strip()
    )
    data = None
    sources = []
    provider_failed = False

    if os.getenv("KGD_PORTAL_TOKEN", "").strip():
        try:
            data = _lookup_kgd_taxpayer(identifier)
            if data:
                sources.append("kgd")
        except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
            provider_failed = True
            current_app.logger.exception("KGD taxpayer lookup failed")

    # Enrich KGD data with address/phone from the previously configured
    # registry. Values returned by KGD always keep priority.
    if os.getenv("CLIENT_LOOKUP_API_URL", "").strip():
        try:
            registry_data = _lookup_configured_provider(identifier)
            if registry_data:
                data = _merge_lookup_data(data, registry_data)
                sources.append("registry")
        except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
            provider_failed = True
            current_app.logger.exception("Configured client registry lookup failed")

    # eGov is queried only while the address is still missing, which avoids an
    # unnecessary external request when the previous registry filled it.
    if (
        os.getenv("EGOV_OPEN_DATA_API_KEY", "").strip()
        and not (data or {}).get("address")
    ):
        try:
            egov_data = _lookup_egov_open_data(identifier)
            if egov_data:
                data = _merge_lookup_data(data, egov_data)
                sources.append("egov")
        except (HTTPError, URLError, TimeoutError, ValueError, json.JSONDecodeError):
            provider_failed = True
            current_app.logger.exception("eGov client registry lookup failed")

    if provider_failed and not data:
        return jsonify({
            "found": False,
            "message": "Внешние справочники временно недоступны. Данные можно заполнить вручную",
        }), 502

    if data:
        data_source = "+".join(sources)
        enriched = len(sources) > 1
        return jsonify({
            "found": True,
            "source": data_source,
            "data": data,
            "message": (
                "Данные найдены в КГД и дополнены из справочника"
                if enriched and "kgd" in sources
                else "Данные найдены в КГД"
                if data_source == "kgd"
                else "Данные найдены в справочнике"
            ),
        })

    message = "По этому ИИН/БИН данные не найдены"
    if not provider_configured:
        message = "В базе клиента нет. Внешний справочник ещё не подключён"
    return jsonify({"found": False, "message": message})

# ================== ИМПОРТ / ЭКСПОРТ КЛИЕНТОВ ==================

_CLIENT_HEADERS = [
    "ФИО / Наименование", "Телефон", "ИИН / БИН", "Компания",
    "Адрес", "Статус", "Категория", "Оплата", "Комментарий"
]

_CLIENT_ALIASES = {
    "фио / наименование": "full_name", "фио": "full_name",
    "наименование": "full_name", "клиент": "full_name",
    "телефон": "phone", "номер телефона": "phone",
    "иин / бин": "iin", "иин": "iin", "бин": "iin",
    "компания": "company_name", "название компании": "company_name",
    "адрес": "address", "статус": "status", "категория": "category",
    "оплата": "payment", "комментарий": "comment",
}


def _client_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _style_clients_excel(ws):
    fill = PatternFill("solid", fgColor="252B3A")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    widths = [32, 20, 18, 28, 36, 16, 18, 18, 42]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.auto_filter.ref = ws.dimensions


@clients_bp.route("/clients/export.xlsx")
def export_clients_xlsx():
    company_id = _company_id()
    if not company_id:
        return "Компания не выбрана", 403

    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT full_name, phone, iin, company_name, address,
                   status, category, payment, comment
            FROM clients
            WHERE company_id = %s AND COALESCE(is_deleted, FALSE) = FALSE
            ORDER BY full_name
        """, (company_id,))
        rows = cur.fetchall()
    finally:
        pool.putconn(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(_CLIENT_HEADERS)
    for row in rows:
        ws.append([row.get(key) or "" for key in (
            "full_name", "phone", "iin", "company_name", "address",
            "status", "category", "payment", "comment"
        )])
    _style_clients_excel(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"nika_clients_company_{company_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@clients_bp.route("/clients/import-template.xlsx")
def clients_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(_CLIENT_HEADERS)
    ws.append([
        "ТОО Пример", "+7 777 000 00 00", "123456789012",
        "ТОО Пример", "г. Усть-Каменогорск", "Новый",
        "Корпоративный", "Не оплачено", "Пример — строку можно удалить"
    ])
    _style_clients_excel(ws)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="nika_clients_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@clients_bp.route("/clients/import", methods=["POST"])
def import_clients_xlsx():
    company_id = _company_id()
    if not company_id:
        return jsonify({"success": False, "message": "Компания не выбрана"}), 403

    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"success": False, "message": "Выберите Excel-файл"}), 400
    if not upload.filename.lower().endswith(".xlsx"):
        return jsonify({"success": False, "message": "Поддерживается формат .xlsx"}), 400

    duplicate_mode = request.form.get("duplicate_mode", "skip")
    if duplicate_mode not in ("skip", "update"):
        duplicate_mode = "skip"

    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"success": False, "message": "Не удалось открыть Excel-файл"}), 400

    headers = [_client_text(cell.value).lower() for cell in ws[1]]
    header_map = {}
    for column_index, header in enumerate(headers, start=1):
        field = _CLIENT_ALIASES.get(header)
        if field:
            header_map[field] = column_index

    if "full_name" not in header_map:
        return jsonify({
            "success": False,
            "message": "В файле нет обязательной колонки «ФИО / Наименование»"
        }), 400

    def value(row, field):
        column = header_map.get(field)
        return _client_text(row[column - 1]) if column else ""

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    conn = get_db()
    try:
        cur = conn.cursor()
        for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            full_name = value(row, "full_name")
            if not full_name:
                continue
            try:
                phone = value(row, "phone")
                iin = value(row, "iin")
                company_name = value(row, "company_name")
                address = value(row, "address")
                status = value(row, "status") or "Новый"
                category = value(row, "category")
                payment = value(row, "payment") or "Не оплачено"
                comment = value(row, "comment")

                existing = None
                if iin:
                    cur.execute("""
                        SELECT id FROM clients
                        WHERE company_id=%s AND iin=%s LIMIT 1
                    """, (company_id, iin))
                    existing = cur.fetchone()
                if not existing and phone:
                    cur.execute("""
                        SELECT id FROM clients
                        WHERE company_id=%s AND phone=%s LIMIT 1
                    """, (company_id, phone))
                    existing = cur.fetchone()

                if existing and duplicate_mode == "skip":
                    stats["skipped"] += 1
                    continue

                if existing:
                    cur.execute("""
                        UPDATE clients SET
                            full_name=%s, phone=%s, iin=%s, company_name=%s,
                            address=%s, status=%s, category=%s, payment=%s,
                            comment=%s, is_deleted=FALSE
                        WHERE id=%s AND company_id=%s
                    """, (
                        full_name, phone, iin, company_name, address, status,
                        category, payment, comment, existing["id"], company_id
                    ))
                    stats["updated"] += 1
                else:
                    cur.execute("""
                        INSERT INTO clients (
                            full_name, phone, iin, company_name, address,
                            status, category, payment, comment, created_at,
                            company_id, is_deleted
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,FALSE)
                    """, (
                        full_name, phone, iin, company_name, address, status,
                        category, payment, comment, now_kz(), company_id
                    ))
                    stats["created"] += 1
            except Exception as row_error:
                stats["errors"].append({"row": excel_row, "message": str(row_error)[:180]})
                if len(stats["errors"]) >= 50:
                    break

        conn.commit()
        return jsonify({"success": True, **stats})
    except Exception as error:
        conn.rollback()
        return jsonify({"success": False, "message": str(error)}), 500
    finally:
        pool.putconn(conn)
