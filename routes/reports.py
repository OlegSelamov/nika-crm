from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Blueprint,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models import get_db, pool
from utils.timezone import now_kz


reports_bp = Blueprint("reports", __name__)

REPORT_TITLES = {
    "sales": "Отчёт по продажам",
    "profit": "Отчёт по прибыли",
    "stock": "Отчёт по складу",
    "clients": "Отчёт по клиентам",
}


def _require_company():
    if not session.get("user_id"):
        return None
    return session.get("company_id")


def _parse_period():
    today = now_kz().date()
    default_from = today.replace(day=1)

    raw_from = request.args.get("date_from")
    raw_to = request.args.get("date_to")

    try:
        date_from = datetime.strptime(raw_from, "%Y-%m-%d").date() if raw_from else default_from
        date_to = datetime.strptime(raw_to, "%Y-%m-%d").date() if raw_to else today
    except ValueError:
        date_from, date_to = default_from, today

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return date_from, date_to


def _money(value):
    return float(value or 0)


def _get_summary(cur, company_id, date_from, date_to):
    cur.execute("""
        SELECT
            COUNT(*) AS sales_count,
            COALESCE(SUM(total_amount), 0) AS revenue,
            COALESCE(SUM(cash_amount), 0) AS cash,
            COALESCE(SUM(card_amount), 0) AS card,
            COALESCE(SUM(kaspi_amount), 0) AS kaspi,
            COALESCE(AVG(total_amount), 0) AS average_check
        FROM sales
        WHERE company_id = %s
          AND status = 'Оплачено'
          AND DATE(created_at) BETWEEN %s AND %s
    """, (company_id, date_from, date_to))
    sales = cur.fetchone()

    cur.execute("""
        SELECT COALESCE(SUM(si.profit), 0) AS profit
        FROM sale_items si
        JOIN sales s ON s.id = si.sale_id
        WHERE s.company_id = %s
          AND s.status = 'Оплачено'
          AND DATE(s.created_at) BETWEEN %s AND %s
    """, (company_id, date_from, date_to))
    profit = cur.fetchone()["profit"] or 0

    expenses = 0
    try:
        cur.execute("""
            SELECT COALESCE(SUM(amount), 0) AS expenses
            FROM expenses
            WHERE company_id = %s
              AND expense_date BETWEEN %s AND %s
        """, (company_id, date_from, date_to))
        expenses = cur.fetchone()["expenses"] or 0
    except Exception:
        cur.connection.rollback()

    return {
        "sales_count": sales["sales_count"] or 0,
        "revenue": sales["revenue"] or 0,
        "cash": sales["cash"] or 0,
        "card": sales["card"] or 0,
        "kaspi": sales["kaspi"] or 0,
        "average_check": sales["average_check"] or 0,
        "gross_profit": profit,
        "expenses": expenses,
        "net_profit": (profit or 0) - (expenses or 0),
    }


def _sales_report(cur, company_id, date_from, date_to):
    cur.execute("""
        SELECT
            s.id,
            s.sale_number,
            s.created_at,
            COALESCE(c.company_name, c.full_name, 'Без клиента') AS client,
            s.total_amount,
            s.cash_amount,
            s.card_amount,
            s.kaspi_amount,
            s.status
        FROM sales s
        LEFT JOIN clients c ON c.id = s.client_id
        WHERE s.company_id = %s
          AND DATE(s.created_at) BETWEEN %s AND %s
        ORDER BY s.created_at DESC, s.id DESC
        LIMIT 1000
    """, (company_id, date_from, date_to))

    rows = []
    for row in cur.fetchall():
        payment = []
        if _money(row["cash_amount"]) > 0:
            payment.append("Наличные")
        if _money(row["card_amount"]) > 0:
            payment.append("Карта")
        if _money(row["kaspi_amount"]) > 0:
            payment.append("Kaspi")

        rows.append({
            "date": row["created_at"].strftime("%d.%m.%Y %H:%M"),
            "number": row["sale_number"] or row["id"],
            "client": row["client"],
            "payment": ", ".join(payment) or "Не указан",
            "amount": row["total_amount"] or 0,
            "status": row["status"],
        })

    return {
        "columns": [
            ("date", "Дата"),
            ("number", "№ чека"),
            ("client", "Клиент"),
            ("payment", "Оплата"),
            ("amount", "Сумма"),
            ("status", "Статус"),
        ],
        "rows": rows,
    }


def _profit_report(cur, company_id, date_from, date_to):
    cur.execute("""
        SELECT
            DATE(s.created_at) AS report_date,
            COUNT(DISTINCT s.id) AS sales_count,
            COALESCE(SUM(si.total), 0) AS revenue,
            COALESCE(SUM(si.profit), 0) AS gross_profit
        FROM sales s
        JOIN sale_items si ON si.sale_id = s.id
        WHERE s.company_id = %s
          AND s.status = 'Оплачено'
          AND DATE(s.created_at) BETWEEN %s AND %s
        GROUP BY DATE(s.created_at)
        ORDER BY report_date DESC
    """, (company_id, date_from, date_to))

    rows = []
    for row in cur.fetchall():
        rows.append({
            "date": row["report_date"].strftime("%d.%m.%Y"),
            "sales_count": row["sales_count"],
            "revenue": row["revenue"] or 0,
            "profit": row["gross_profit"] or 0,
            "margin": (
                round(_money(row["gross_profit"]) / _money(row["revenue"]) * 100, 1)
                if _money(row["revenue"]) else 0
            ),
        })

    return {
        "columns": [
            ("date", "Дата"),
            ("sales_count", "Продаж"),
            ("revenue", "Выручка"),
            ("profit", "Валовая прибыль"),
            ("margin", "Маржа, %"),
        ],
        "rows": rows,
    }


def _stock_report(cur, company_id, date_from, date_to):
    cur.execute("""
        SELECT
            i.id,
            i.name,
            COALESCE(i.category, 'Без категории') AS category,
            COALESCE(i.quantity, 0) AS quantity,
            COALESCE(i.purchase_price, 0) AS purchase_price,
            COALESCE(i.retail_price, i.price, 0) AS retail_price,
            COALESCE(i.quantity, 0) * COALESCE(i.purchase_price, 0) AS stock_cost
        FROM items i
        WHERE i.company_id = %s
        ORDER BY i.name
        LIMIT 2000
    """, (company_id,))

    rows = []
    for row in cur.fetchall():
        rows.append({
            "name": row["name"],
            "category": row["category"],
            "quantity": row["quantity"],
            "purchase_price": row["purchase_price"],
            "retail_price": row["retail_price"],
            "stock_cost": row["stock_cost"],
        })

    return {
        "columns": [
            ("name", "Товар"),
            ("category", "Категория"),
            ("quantity", "Остаток"),
            ("purchase_price", "Закупочная цена"),
            ("retail_price", "Розничная цена"),
            ("stock_cost", "Стоимость остатка"),
        ],
        "rows": rows,
    }


def _clients_report(cur, company_id, date_from, date_to):
    cur.execute("""
        SELECT
            c.id,
            COALESCE(c.company_name, c.full_name, 'Без имени') AS client,
            c.phone,
            COUNT(s.id) FILTER (WHERE s.status = 'Оплачено') AS purchases,
            COALESCE(SUM(s.total_amount) FILTER (WHERE s.status = 'Оплачено'), 0) AS total,
            MAX(s.created_at) FILTER (WHERE s.status = 'Оплачено') AS last_purchase
        FROM clients c
        LEFT JOIN sales s
          ON s.client_id = c.id
         AND s.company_id = %s
         AND DATE(s.created_at) BETWEEN %s AND %s
        WHERE c.company_id = %s
          AND COALESCE(c.is_deleted, FALSE) = FALSE
        GROUP BY c.id, c.company_name, c.full_name, c.phone
        ORDER BY total DESC, client
        LIMIT 1000
    """, (company_id, date_from, date_to, company_id))

    rows = []
    for row in cur.fetchall():
        rows.append({
            "client": row["client"],
            "phone": row["phone"] or "—",
            "purchases": row["purchases"] or 0,
            "total": row["total"] or 0,
            "last_purchase": (
                row["last_purchase"].strftime("%d.%m.%Y")
                if row["last_purchase"] else "—"
            ),
        })

    return {
        "columns": [
            ("client", "Клиент"),
            ("phone", "Телефон"),
            ("purchases", "Покупок"),
            ("total", "Сумма покупок"),
            ("last_purchase", "Последняя покупка"),
        ],
        "rows": rows,
    }


def _build_report(cur, report_type, company_id, date_from, date_to):
    builders = {
        "sales": _sales_report,
        "profit": _profit_report,
        "stock": _stock_report,
        "clients": _clients_report,
    }
    return builders.get(report_type, _sales_report)(
        cur, company_id, date_from, date_to
    )


@reports_bp.route("/reports")
def reports():
    if not session.get("user_id"):
        return redirect("/login")

    company_id = _require_company()
    if not company_id:
        return "Активная компания не выбрана", 400

    date_from, date_to = _parse_period()
    conn = get_db()
    cur = conn.cursor()

    try:
        summary = _get_summary(cur, company_id, date_from, date_to)
        return render_template(
            "reports.html",
            summary=summary,
            date_from=date_from.isoformat(),
            date_to=date_to.isoformat(),
            active_report="sales",
            report_title=REPORT_TITLES["sales"],
            report=None,
            print_mode=False,
        )
    finally:
        cur.close()
        pool.putconn(conn)


@reports_bp.route("/reports/data")
def reports_data():
    if not session.get("user_id"):
        return jsonify({"success": False, "error": "Требуется авторизация"}), 401

    company_id = _require_company()
    report_type = request.args.get("type", "sales")
    date_from, date_to = _parse_period()

    conn = get_db()
    cur = conn.cursor()

    try:
        report = _build_report(
            cur, report_type, company_id, date_from, date_to
        )
        return jsonify({
            "success": True,
            "title": REPORT_TITLES.get(report_type, "Отчёт"),
            "columns": [
                {"key": key, "label": label}
                for key, label in report["columns"]
            ],
            "rows": report["rows"],
        })
    finally:
        cur.close()
        pool.putconn(conn)


@reports_bp.route("/reports/export.xlsx")
def export_reports_excel():
    if not session.get("user_id"):
        return redirect("/login")

    company_id = _require_company()
    report_type = request.args.get("type", "sales")
    date_from, date_to = _parse_period()

    conn = get_db()
    cur = conn.cursor()

    try:
        report = _build_report(
            cur, report_type, company_id, date_from, date_to
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Отчёт"

        title = REPORT_TITLES.get(report_type, "Отчёт")
        sheet.append([title])
        sheet.append([f"Период: {date_from:%d.%m.%Y} — {date_to:%d.%m.%Y}"])
        sheet.append([])

        headers = [label for _, label in report["columns"]]
        sheet.append(headers)

        header_row = sheet.max_row
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="7257FF")
            cell.alignment = Alignment(horizontal="center")

        for item in report["rows"]:
            sheet.append([item.get(key, "") for key, _ in report["columns"]])

        for column in sheet.columns:
            max_length = max(
                len(str(cell.value or "")) for cell in column
            )
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(max_length + 2, 12), 42
            )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = (
            f"{report_type}_{date_from.isoformat()}_{date_to.isoformat()}.xlsx"
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
    finally:
        cur.close()
        pool.putconn(conn)


@reports_bp.route("/reports/print")
def print_report():
    if not session.get("user_id"):
        return redirect("/login")

    company_id = _require_company()
    report_type = request.args.get("type", "sales")
    date_from, date_to = _parse_period()

    conn = get_db()
    cur = conn.cursor()

    try:
        summary = _get_summary(cur, company_id, date_from, date_to)
        report = _build_report(
            cur, report_type, company_id, date_from, date_to
        )

        return render_template(
            "reports.html",
            summary=summary,
            date_from=date_from.isoformat(),
            date_to=date_to.isoformat(),
            active_report=report_type,
            report_title=REPORT_TITLES.get(report_type, "Отчёт"),
            report=report,
            print_mode=True,
        )
    finally:
        cur.close()
        pool.putconn(conn)